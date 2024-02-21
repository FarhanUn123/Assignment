from selenium import webdriver
import openpyxl
import time
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By

workbook = openpyxl.load_workbook('C:/Users/ufarh/Downloads/Excel.xlsx')

for sheetname in workbook.sheetnames:
    ws = workbook[sheetname]

    driver = webdriver.Chrome()
    driver.maximize_window()

    driver.get("https://www.google.com/")
    for i in range(3,13):
        search=driver.find_element("name", "q")
        search.send_keys(ws.cell(row=i, column=3).value)
        time.sleep(3)
        autocomplete_options= driver.find_element(By.XPATH, "//*[contains(@class, 'erkvQe')]")

        a= autocomplete_options.text.split('\n')

        shortest= min(a, key=len)
        longest= max(a, key=len)

        ws.cell(row=i, column=4, value=longest)
        ws.cell(row=i, column=5, value=shortest)

        search.clear()

workbook.save('C:/Users/ufarh/Downloads/Excel2.xlsx')
driver.close()
workbook.close()
print("Successfully completed")